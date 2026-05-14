import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import json
from collections import Counter
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Supabase
try:
    from supabase import create_client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

def is_cloud():
    try:
        return "supabase_url" in st.secrets
    except:
        return False

# ─────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────
st.set_page_config(
    page_title="Torre de Control · FL Servital",
    layout="wide",
    page_icon="🔧",
    initial_sidebar_state="collapsed"
)

# CSS personalizado — limpio, industrial, fácil de leer
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Header principal */
.main-header {
    background: #1a1a2e;
    color: white;
    padding: 16px 24px;
    border-radius: 12px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 12px;
}
.main-header h1 { margin: 0; font-size: 22px; font-weight: 700; }
.main-header p  { margin: 0; font-size: 13px; opacity: 0.6; }

/* Tarjetas de operario */
.operario-card {
    background: white;
    border-radius: 12px;
    padding: 16px;
    border: 2px solid #e8e8e8;
    margin-bottom: 12px;
    transition: border-color 0.2s;
}
.operario-name {
    font-size: 13px;
    font-weight: 700;
    color: #666;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 8px;
}
.operario-task {
    font-size: 15px;
    font-weight: 600;
    color: #1a1a2e;
}
.operario-sub {
    font-size: 12px;
    color: #888;
    font-family: 'DM Mono', monospace;
}

/* Badges de estado */
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.3px;
}
.badge-des  { background: #FFF3CD; color: #856404; }
.badge-pin  { background: #CCE5FF; color: #004085; }
.badge-arm  { background: #FFE0CC; color: #7a3a00; }
.badge-pul  { background: #E8D5FF; color: #4a0080; }
.badge-term { background: #D4EDDA; color: #155724; }
.badge-lav  { background: #D1ECF1; color: #0c5460; }
.badge-ok   { background: #D4EDDA; color: #155724; }
.badge-none { background: #f0f0f0; color: #888; }

/* Barra de progreso */
.progress-container {
    background: #f0f0f0;
    border-radius: 6px;
    height: 8px;
    overflow: hidden;
    margin: 6px 0;
}
.progress-bar {
    height: 100%;
    border-radius: 6px;
    transition: width 0.3s;
}

/* Fila del panel */
.panel-row {
    background: white;
    border-radius: 10px;
    padding: 12px 16px;
    margin-bottom: 8px;
    border: 1px solid #e8e8e8;
    display: flex;
    align-items: center;
    gap: 12px;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    background: #f5f5f5;
    padding: 4px;
    border-radius: 10px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    font-weight: 600;
    font-size: 14px;
}

/* Botón avanzar */
.stButton button {
    border-radius: 8px;
    font-weight: 600;
}

/* Métrica */
.metric-box {
    background: white;
    border-radius: 10px;
    padding: 16px;
    text-align: center;
    border: 1px solid #e8e8e8;
}
.metric-num { font-size: 32px; font-weight: 700; color: #1a1a2e; }
.metric-label { font-size: 12px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }

/* Alerta urgente */
.alerta-urgente {
    background: #fff5f5;
    border-left: 4px solid #e53e3e;
    padding: 10px 14px;
    border-radius: 6px;
    margin-bottom: 8px;
    font-size: 13px;
}

div[data-testid="stHorizontalBlock"] { gap: 12px; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────
ETAPAS = ['desabolladura', 'pintura', 'armado', 'pulido', 'terminaciones', 'lavado']
ETAPA_LABEL = {
    'desabolladura': 'Desabolladura',
    'pintura':       'Pintura',
    'armado':        'Armado',
    'pulido':        'Pulido',
    'terminaciones': 'Terminaciones',
    'lavado':        'Lavado',
}
ETAPA_BADGE = {
    'desabolladura': 'des', 'pintura': 'pin', 'armado': 'arm',
    'pulido': 'pul', 'terminaciones': 'term', 'lavado': 'lav',
}
ETAPA_COLOR = {
    'desabolladura': '#856404',
    'pintura':       '#004085',
    'armado':        '#7a3a00',
    'pulido':        '#4a0080',
    'terminaciones': '#155724',
    'lavado':        '#0c5460',
}
ETAPA_PROGRESS = {
    'desabolladura': 1, 'pintura': 2, 'armado': 3,
    'pulido': 4, 'terminaciones': 5, 'lavado': 6,
}
SIGUIENTE_ETAPA = {
    'desabolladura': 'pintura',
    'pintura':       'armado',
    'armado':        'pulido',
    'pulido':        'terminaciones',
    'terminaciones': 'lavado',
    'lavado':        None,
}
ESTADO_MAP = {
    'DES': 'desabolladura', 'PIN': 'pintura', 'ARM': 'armado',
    'PUL': 'pulido', 'LAV': 'lavado', 'OK.': 'listo',
}
DES_MAP = {'RO': 'Rojas', 'GO': 'González', 'CA': 'Carvajal', 'MEC': 'Externo', 'EXT': 'Externo'}
PIN_MAP = {'AS': 'Percy', 'HU': 'Huaiquifil', 'M': 'Percy', 'ME': 'Huaiquifil', 'PE': 'Percy'}

DESABOLLADORES = ['Rojas', 'González', 'Carvajal']
PINTORES       = ['Percy', 'Huaiquifil']

DURACIONES = {
    'LEVE':    {'desabolladura': 1, 'pintura': 1, 'armado': 1, 'pulido': 1, 'terminaciones': 1, 'lavado': 1},
    'MEDIANO': {'desabolladura': 2, 'pintura': 2, 'armado': 1, 'pulido': 1, 'terminaciones': 1, 'lavado': 1},
    'GRAVE':   {'desabolladura': 5, 'pintura': 5, 'armado': 3, 'pulido': 2, 'terminaciones': 2, 'lavado': 1},
}

FERIADOS = [
    date(2026, 1, 1), date(2026, 4, 3), date(2026, 5, 1),
    date(2026, 9, 18), date(2026, 9, 19), date(2026, 12, 25)
]

hoy = date.today()

# ── Pesos de asignación por defecto (% por operario y tipo de daño)
PESOS_DEFAULT = {
    'desabolladores': {
        'LEVE':    {'Rojas': 34, 'González': 33, 'Carvajal': 33},
        'MEDIANO': {'Rojas': 34, 'González': 33, 'Carvajal': 33},
        'GRAVE':   {'Rojas': 34, 'González': 33, 'Carvajal': 33},
    },
    'pintores': {
        'LEVE':    {'Percy': 50, 'Huaiquifil': 50},
        'MEDIANO': {'Percy': 50, 'Huaiquifil': 50},
        'GRAVE':   {'Percy': 50, 'Huaiquifil': 50},
    }
}

def get_pesos():
    if 'pesos_asignacion' not in st.session_state:
        # Intentar cargar desde Supabase o SQLite
        try:
            if supa:
                resp = supa.table("configuracion").select("valor").eq("clave", "pesos_asignacion").execute()
                if resp.data:
                    st.session_state.pesos_asignacion = json.loads(resp.data[0]['valor'])
                else:
                    st.session_state.pesos_asignacion = PESOS_DEFAULT
            else:
                c = conn.cursor()
                c.execute("SELECT valor FROM configuracion WHERE clave='pesos_asignacion'")
                row = c.fetchone()
                if row:
                    st.session_state.pesos_asignacion = json.loads(row[0])
                else:
                    st.session_state.pesos_asignacion = PESOS_DEFAULT
        except:
            st.session_state.pesos_asignacion = PESOS_DEFAULT
    return st.session_state.pesos_asignacion

def save_pesos(pesos):
    valor = json.dumps(pesos)
    try:
        if supa:
            supa.table("configuracion").upsert({"clave": "pesos_asignacion", "valor": valor}).execute()
        else:
            c = conn.cursor()
            c.execute("CREATE TABLE IF NOT EXISTS configuracion (clave TEXT PRIMARY KEY, valor TEXT)")
            c.execute("INSERT OR REPLACE INTO configuracion VALUES ('pesos_asignacion', ?)", (valor,))
            conn.commit()
        st.session_state.pesos_asignacion = pesos
    except Exception as e:
        st.error(f"Error guardando pesos: {e}")

# Posiciones del taller
POSICIONES_CARROCERIA = [f"Carr{i}" for i in range(1,17)]
POSICIONES_ESP_PINTURA = [f"P{i}" for i in range(1,16)]
POSICIONES_ESP_ARMADO  = [f"Est{i}" for i in range(11,19)]
POSICIONES_ESP_PULIDO  = [f"Est{i}" for i in range(19,31)]
POSICIONES_LAVADO      = ["Lav1","Lav2","Lav3"]
POSICIONES_PRE_ENTREGA = ["PE1","PE2","PE3","PE4"]
POSICIONES_PINTURA     = [f"Pint{i}" for i in range(1,17)]
POSICIONES_HORNOS      = ["H1","H2","H3"]
TODAS_POSICIONES = (["— Sin asignar —"] +
    POSICIONES_CARROCERIA + POSICIONES_ESP_PINTURA +
    POSICIONES_ESP_ARMADO + POSICIONES_ESP_PULIDO +
    POSICIONES_LAVADO + POSICIONES_PRE_ENTREGA +
    POSICIONES_PINTURA + POSICIONES_HORNOS)

# ─────────────────────────────────────────────────────
# BASE DE DATOS (SQLite)
# ─────────────────────────────────────────────────────
import sqlite3

@st.cache_resource
def get_db():
    conn = sqlite3.connect("taller_v2.db", check_same_thread=False)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS ots (
        ot TEXT PRIMARY KEY,
        nombre TEXT, cia TEXT, ase TEXT,
        modelo TEXT, color TEXT, patente TEXT,
        dano TEXT,
        desabollador TEXT, pintor TEXT,
        etapa_actual TEXT,
        fecha_ingreso TEXT, fecha_entrega TEXT,
        comentario TEXT, repuestos TEXT, telefono TEXT,
        programacion_json TEXT,
        entregado INTEGER DEFAULT 0,
        posicion TEXT
    )''')
    try: c.execute('ALTER TABLE ots ADD COLUMN posicion TEXT')
    except: pass
    conn.commit()
    return conn

conn = get_db()

# Inicializar Supabase
def _init_supa():
    if not SUPABASE_AVAILABLE: return None
    try:
        url = st.secrets.get("supabase_url","")
        key = st.secrets.get("supabase_key","")
        if url and key:
            client = create_client(url, key)
            return client
    except Exception as e:
        st.warning(f"No se pudo conectar con Supabase: {e}")
    return None

supa = _init_supa()

def save_ot(data):
    prog = json.dumps(data.get('programacion', {}))
    if supa:
        supa.table("ots").upsert({
            'ot': data['ot'], 'nombre': data.get('nombre',''),
            'cia': data.get('cia',''), 'ase': data.get('ase',''),
            'modelo': data['modelo'], 'color': data.get('color',''),
            'patente': data['patente'], 'dano': data.get('dano','MEDIANO'),
            'desabollador': data['desabollador'], 'pintor': data['pintor'],
            'etapa_actual': data['etapa_actual'],
            'fecha_ingreso': data.get('fecha_ingreso',''),
            'fecha_entrega': data.get('fecha_entrega',''),
            'comentario': data.get('comentario',''),
            'repuestos': data.get('repuestos',''),
            'telefono': data.get('telefono',''),
            'programacion_json': prog,
            'entregado': 0, 'posicion': data.get('posicion','')
        }).execute()
    else:
        c = conn.cursor()
        c.execute('''INSERT OR REPLACE INTO ots VALUES
            (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            data['ot'], data.get('nombre',''), data.get('cia',''), data.get('ase',''),
            data['modelo'], data.get('color',''), data['patente'],
            data.get('dano','MEDIANO'), data['desabollador'], data['pintor'],
            data['etapa_actual'], data.get('fecha_ingreso',''), data.get('fecha_entrega',''),
            data.get('comentario',''), data.get('repuestos',''), data.get('telefono',''),
            prog, 0, data.get('posicion','')
        ))
        conn.commit()

def load_ots():
    if supa:
        try:
            resp = supa.table("ots").select("*").eq("entregado", 0).order("fecha_ingreso").execute()
            rows = []
            for r in resp.data:
                try: r['programacion'] = json.loads(r.get('programacion_json') or '{}')
                except: r['programacion'] = {}
                rows.append(r)
            return rows
        except Exception as e:
            st.error(f"⚠️ Error conectando con Supabase: {e}. Reintentando...")
            # Reintentar con nueva conexión
            try:
                new_supa = create_client(st.secrets["supabase_url"], st.secrets["supabase_key"])
                resp = new_supa.table("ots").select("*").eq("entregado", 0).order("fecha_ingreso").execute()
                rows = []
                for r in resp.data:
                    try: r['programacion'] = json.loads(r.get('programacion_json') or '{}')
                    except: r['programacion'] = {}
                    rows.append(r)
                return rows
            except Exception as e2:
                st.error(f"❌ No se pudo conectar con Supabase: {e2}")
                return []
    else:
        c = conn.cursor()
        c.execute("SELECT * FROM ots WHERE entregado=0 ORDER BY fecha_ingreso")
        cols = ['ot','nombre','cia','ase','modelo','color','patente','dano',
                'desabollador','pintor','etapa_actual','fecha_ingreso','fecha_entrega',
                'comentario','repuestos','telefono','programacion_json','entregado','posicion']
        rows = []
        for row in c.fetchall():
            d = dict(zip(cols, row))
            try: d['programacion'] = json.loads(d['programacion_json'])
            except: d['programacion'] = {}
            rows.append(d)
        return rows

def marcar_entregado(ot):
    if supa:
        supa.table("ots").update({"entregado": 1}).eq("ot", ot).execute()
    else:
        c = conn.cursor()
        c.execute("UPDATE ots SET entregado=1 WHERE ot=?", (ot,))
        conn.commit()

def avanzar_etapa(ot, nueva_etapa):
    if supa:
        supa.table("ots").update({"etapa_actual": nueva_etapa}).eq("ot", ot).execute()
    else:
        c = conn.cursor()
        c.execute("UPDATE ots SET etapa_actual=? WHERE ot=?", (nueva_etapa, ot))
        conn.commit()

def update_ot_field(ot, field, value):
    if supa:
        supa.table("ots").update({field: value}).eq("ot", ot).execute()
    else:
        c = conn.cursor()
        c.execute(f"UPDATE ots SET {field}=? WHERE ot=?", (value, ot))
        conn.commit()

# ─────────────────────────────────────────────────────
# FUNCIONES DE PROGRAMACIÓN
# ─────────────────────────────────────────────────────
def es_laboral(d):
    return d.weekday() < 5 and d not in FERIADOS

def dias_laborales_desde(inicio, n):
    """Retorna la fecha fin después de n días laborales desde inicio"""
    d = inicio
    count = 0
    while count < n:
        if es_laboral(d):
            count += 1
        if count < n:
            d += timedelta(days=1)
    return d

def programar_ot(desabollador, pintor, dano, fecha_inicio, ots_existentes, etapa_desde=None):
    """Calcula fechas de cada etapa respetando ocupación de operarios.
    etapa_desde: si se indica, las etapas anteriores se marcan como completadas."""
    dano_key = dano.upper() if dano else 'MEDIANO'
    durs = DURACIONES.get(dano_key, DURACIONES['MEDIANO'])
    ETAPA_IDX = {e: i for i, e in enumerate(ETAPAS)}
    idx_desde = ETAPA_IDX.get(etapa_desde, 0) if etapa_desde else 0

    # Construir calendarios de ocupación
    cal_des = {}  # {fecha: n_ots}
    cal_pin = {}
    for ot in ots_existentes:
        prog = ot.get('programacion', {})
        des = ot.get('desabollador', '')
        pin = ot.get('pintor', '')
        for etapa, info in prog.items():
            if not info: continue
            try:
                fi = datetime.strptime(info['inicio'], '%Y-%m-%d').date()
                ff = datetime.strptime(info['fin'], '%Y-%m-%d').date()
            except: continue
            d = fi
            while d <= ff:
                if es_laboral(d):
                    if etapa in ('desabolladura', 'armado') and des == desabollador:
                        cal_des[d] = cal_des.get(d, 0) + 1
                    if etapa in ('pintura', 'pulido') and pin == pintor:
                        cap = 2 if 'Huaiquifil' in pin else 1
                        cal_pin[d] = cal_pin.get(d, 0) + 1
                d += timedelta(days=1)

    def buscar_inicio_libre(cal, desde, dur, capacidad=1):
        d = max(desde, hoy)
        while True:
            ok = True
            temp = d
            dias = 0
            while dias < dur:
                if es_laboral(temp):
                    if cal.get(temp, 0) >= capacidad:
                        ok = False
                        break
                    dias += 1
                temp += timedelta(days=1)
            if ok:
                return d
            d += timedelta(days=1)

    cap_pin = 2 if 'Huaiquifil' in pintor else 1
    prog = {}
    auto_desde = fecha_inicio

    for etapa in ETAPAS:
        dur = durs[etapa]
        idx_etapa = ETAPA_IDX.get(etapa, 0)
        # Si la etapa ya fue completada, marcarla como tal sin bloquear calendario
        if idx_etapa < idx_desde:
            ayer = hoy - timedelta(days=1)
            prog[etapa] = {
                'inicio': ayer.strftime('%Y-%m-%d'),
                'fin': ayer.strftime('%Y-%m-%d'),
                'completado': True
            }
            continue
        if etapa in ('terminaciones', 'lavado'):
            inicio = auto_desde
            fin = dias_laborales_desde(inicio, dur)
            prog[etapa] = {'inicio': inicio.strftime('%Y-%m-%d'), 'fin': fin.strftime('%Y-%m-%d')}
        elif etapa in ('desabolladura', 'armado'):
            inicio = buscar_inicio_libre(cal_des, auto_desde, dur, 1)
            fin = dias_laborales_desde(inicio, dur)
            # Bloquear días usados
            d = inicio
            cnt = 0
            while cnt < dur:
                if es_laboral(d):
                    cal_des[d] = cal_des.get(d, 0) + 1
                    cnt += 1
                d += timedelta(days=1)
            prog[etapa] = {'inicio': inicio.strftime('%Y-%m-%d'), 'fin': fin.strftime('%Y-%m-%d')}
        else:  # pintura, pulido
            inicio = buscar_inicio_libre(cal_pin, auto_desde, dur, cap_pin)
            fin = dias_laborales_desde(inicio, dur)
            d = inicio
            cnt = 0
            while cnt < dur:
                if es_laboral(d):
                    cal_pin[d] = cal_pin.get(d, 0) + 1
                    cnt += 1
                d += timedelta(days=1)
            prog[etapa] = {'inicio': inicio.strftime('%Y-%m-%d'), 'fin': fin.strftime('%Y-%m-%d')}

        fin_date = datetime.strptime(prog[etapa]['fin'], '%Y-%m-%d').date()
        auto_desde = fin_date + timedelta(days=1)

    return prog

def mejor_desabollador(ots, dano='MEDIANO'):
    """Elige el desabollador según carga y pesos configurados"""
    pesos = get_pesos()['desabolladores'].get(dano.upper(), PESOS_DEFAULT['desabolladores']['MEDIANO'])
    carga = {d: 0.0 for d in DESABOLLADORES}
    fin_14 = hoy + timedelta(days=14)
    for ot in ots:
        prog = ot.get('programacion', {})
        des = ot.get('desabollador', '')
        if des not in DESABOLLADORES: continue
        for etapa in ['desabolladura', 'armado']:
            info = prog.get(etapa)
            if not info: continue
            try:
                fi = datetime.strptime(info['inicio'], '%Y-%m-%d').date()
                ff = datetime.strptime(info['fin'], '%Y-%m-%d').date()
                if fi <= fin_14:
                    carga[des] += (min(ff, fin_14) - fi).days + 1
            except: pass
    # Ajustar carga según pesos: menor peso = más prioridad
    # Score = carga / peso (a mayor peso, se tolera más carga)
    score = {}
    for d in DESABOLLADORES:
        peso = pesos.get(d, 33)
        score[d] = carga[d] / max(peso, 1)
    return min(score, key=score.get)

def mejor_pintor(ots, dano='MEDIANO'):
    """Elige el pintor según carga, capacidad y pesos configurados"""
    pesos = get_pesos()['pintores'].get(dano.upper(), PESOS_DEFAULT['pintores']['MEDIANO'])
    carga = {p: 0.0 for p in PINTORES}
    fin_14 = hoy + timedelta(days=14)
    for ot in ots:
        prog = ot.get('programacion', {})
        pin = ot.get('pintor', '')
        if pin not in PINTORES: continue
        cap = 2 if pin == 'Huaiquifil' else 1
        for etapa in ['pintura', 'pulido']:
            info = prog.get(etapa)
            if not info: continue
            try:
                fi = datetime.strptime(info['inicio'], '%Y-%m-%d').date()
                ff = datetime.strptime(info['fin'], '%Y-%m-%d').date()
                if fi <= fin_14:
                    carga[pin] += ((min(ff, fin_14) - fi).days + 1) / cap
            except: pass
    score = {}
    for p in PINTORES:
        peso = pesos.get(p, 50)
        score[p] = carga[p] / max(peso, 1)
    return min(score, key=score.get)

# ─────────────────────────────────────────────────────
# HELPERS UI
# ─────────────────────────────────────────────────────
def badge_etapa(etapa):
    if not etapa or etapa == 'listo':
        return "<span class='badge badge-ok'>✅ Listo</span>"
    b = ETAPA_BADGE.get(etapa, 'none')
    label = ETAPA_LABEL.get(etapa, etapa)
    return f"<span class='badge badge-{b}'>{label}</span>"

def progress_pct(etapa):
    if etapa == 'listo': return 100
    return round(ETAPA_PROGRESS.get(etapa, 0) / 6 * 100)

def progress_color(etapa):
    colors = {
        'desabolladura': '#F6AD55', 'pintura': '#63B3ED',
        'armado': '#FC8181', 'pulido': '#B794F4',
        'terminaciones': '#68D391', 'lavado': '#4FD1C5',
        'listo': '#48BB78'
    }
    return colors.get(etapa, '#CBD5E0')

def dias_restantes(fecha_str):
    if not fecha_str: return None
    try:
        f = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        return (f - hoy).days
    except: return None

def format_fecha(fecha_str):
    if not fecha_str: return "—"
    try:
        return datetime.strptime(fecha_str, '%Y-%m-%d').strftime('%d/%m/%y')
    except: return fecha_str

# ─────────────────────────────────────────────────────
# IMPORTAR DESDE PLANILLA (primera vez)
# ─────────────────────────────────────────────────────
def importar_planilla(df_raw):
    importados = 0
    ots_actuales = {ot['ot'] for ot in load_ots()}

    for _, row in df_raw.iterrows():
        try:
            ot_num = str(row.get('O T', '')).strip().replace('.0','')
            if not ot_num or ot_num == 'nan': continue
            ot_id = f"OT-{ot_num}"
            if ot_id in ots_actuales: continue

            estado_raw = str(row.get('EST', '')).strip().upper()
            etapa = ESTADO_MAP.get(estado_raw, 'desabolladura')
            if etapa == 'listo': continue  # ya entregados no se importan

            des_raw = str(row.get('DES', '')).strip().upper()
            pin_raw = str(row.get('PIN', '')).strip().upper()
            des = DES_MAP.get(des_raw, 'Rojas')
            pin = PIN_MAP.get(pin_raw, 'Percy')

            # LLEGA = fecha ingreso al taller (base para programar etapas)
            try:
                fi_raw = pd.to_datetime(row.get('LLEGA'))
                if pd.isna(fi_raw): raise ValueError
                fi = fi_raw.date()
                # Si la fecha ya pasó, programar desde hoy
                if fi < hoy:
                    fi = hoy
                fecha_tentativa = False
            except:
                fi = hoy
                fecha_tentativa = True

            # TALLER = fecha objetivo interna (cuándo debe estar listo)
            try:
                ft_raw = pd.to_datetime(row.get('TALLER'))
                ft_str = ft_raw.strftime('%Y-%m-%d') if not pd.isna(ft_raw) else ''
            except:
                ft_str = ''

            dano_raw = str(row.get('DAÑO', 'MEDIANO')).strip().upper()
            if dano_raw not in ('LEVE','MEDIANO','GRAVE'):
                dano_raw = 'MEDIANO'

            ots_actual = load_ots()
            # Programar solo desde la etapa actual en adelante
            prog = programar_ot(des, pin, dano_raw, fi, ots_actual, etapa_desde=etapa)

            data = {
                'ot': ot_id,
                'nombre': str(row.get('NOMBRE', '') or '').strip(),
                'cia': str(row.get('CIA', '') or '').strip(),
                'ase': str(row.get('ASE', '') or '').strip(),
                'modelo': str(row.get('MODELO', '') or '').strip(),
                'color': str(row.get('COLOR', '') or '').strip(),
                'patente': str(row.get('PATENTE', '') or '').strip(),
                'dano': dano_raw,
                'desabollador': des,
                'pintor': pin,
                'etapa_actual': etapa,
                'fecha_ingreso': fi.strftime('%Y-%m-%d'),
                'fecha_entrega': ft_str,
                'comentario': ('⚠️ Fecha tentativa | ' if fecha_tentativa else '') + str(row.get('COMENTARIO', '') or '').strip(),
                'repuestos': str(row.get('REPUESTOS', '') or '').strip(),
                'telefono': str(row.get('TELEFONO', '') or '').strip(),
                'programacion': prog,
            }
            save_ot(data)
            importados += 1
        except Exception as e:
            continue
    return importados

# ─────────────────────────────────────────────────────
# EXPORTAR A EXCEL
# ─────────────────────────────────────────────────────
def exportar_excel(ots_data):
    ETAPAS_SEQ = ['desabolladura','pintura','armado','pulido','terminaciones','lavado']
    ETAPA_IDX  = {e: i for i, e in enumerate(ETAPAS_SEQ)}
    DES_INV = {'Rojas':'RO','González':'GO','Carvajal':'CA','Externo':'EXT'}
    PIN_INV = {'Percy':'AS','Huaiquifil':'HU'}
    ETAPA_ABREV = {'desabolladura':'DES','pintura':'PIN','armado':'ARM',
                   'pulido':'PUL','terminaciones':'TERM','lavado':'LAV','listo':'OK.'}

    ROJO       = PatternFill("solid", fgColor="C0131A")
    GRIS_OSC   = PatternFill("solid", fgColor="404040")
    GRIS_CLAR  = PatternFill("solid", fgColor="F2F2F2")
    AMARILLO   = PatternFill("solid", fgColor="FFD966")
    VERDE_F    = PatternFill("solid", fgColor="C6EFCE")
    AZUL_F     = PatternFill("solid", fgColor="DDEEFF")
    NARANJO    = PatternFill("solid", fgColor="FCE4D6")
    MORADO     = PatternFill("solid", fgColor="E8D5FF")
    BLANCO_F   = PatternFill("solid", fgColor="FFFFFF")

    borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    izq    = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EN TRABAJO"

    # Título
    ws.merge_cells('A1:Z1')
    ws['A1'] = "PLANILLA D&P — FL SERVITAL"
    ws['A1'].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    ws['A1'].fill = ROJO
    ws['A1'].alignment = centro
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:Z2')
    ws['A2'] = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(ots_data)} vehículos"
    ws['A2'].font = Font(name="Arial", color="FFFFFF", size=8)
    ws['A2'].fill = PatternFill("solid", fgColor="404040")
    ws['A2'].alignment = centro
    ws.row_dimensions[2].height = 14

    headers = [
        ('N°',8),('NOMBRE',22),('CIA',8),('ASE',5),('O T',10),
        ('MODELO',12),('COLOR',8),('PATENTE',10),('DAÑO',9),
        ('DES',6),('PIN',6),
        ('DES✓',6),('PIN✓',6),('ARM✓',6),('PUL✓',6),('LAV✓',6),('OK.',6),
        ('EST',9),('POSICIÓN',10),
        ('COMENTARIO',28),('REPUESTOS',22),
        ('TALLER',10),('CLIENTE',10),('TELÉFONO',13),
    ]
    FILL_H = [PatternFill("solid", fgColor="404040")]*11 +              [AMARILLO,AZUL_F,NARANJO,MORADO,VERDE_F,VERDE_F] +              [PatternFill("solid", fgColor="C0131A")] +              [PatternFill("solid", fgColor="404040")]*6

    for ci, ((h, w), fh) in enumerate(zip(headers, FILL_H), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = fh
        tc = "FFFFFF" if fh.fgColor.rgb in ("FF404040","FFC0131A") else "000000"
        cell.font = Font(name="Arial", bold=True, color=tc, size=9)
        cell.alignment = centro
        cell.border = borde
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 18

    COLOR_EST = {
        'desabolladura': AMARILLO, 'pintura': AZUL_F, 'armado': NARANJO,
        'pulido': MORADO, 'terminaciones': VERDE_F, 'lavado': VERDE_F, 'listo': VERDE_F
    }

    def fdate(s):
        if not s: return ''
        try: return datetime.strptime(s, '%Y-%m-%d').strftime('%d/%m/%Y')
        except: return s

    for rn, ot in enumerate(ots_data, 4):
        etapa = ot.get('etapa_actual','')
        ei = ETAPA_IDX.get(etapa, -1)
        fill_r = COLOR_EST.get(etapa, BLANCO_F)

        def marca(e):
            if etapa == 'listo': return 'X'
            return 'X' if ETAPA_IDX.get(e,-1) < ei else ''

        vals = [
            rn-3,
            ot.get('nombre',''), ot.get('cia',''), ot.get('ase',''),
            ot.get('ot','').replace('OT-',''),
            ot.get('modelo',''), ot.get('color',''), ot.get('patente',''),
            ot.get('dano',''),
            DES_INV.get(ot.get('desabollador',''), ot.get('desabollador','')),
            PIN_INV.get(ot.get('pintor',''), ot.get('pintor','')),
            marca('desabolladura'), marca('pintura'), marca('armado'),
            marca('pulido'), marca('lavado'),
            'X' if etapa == 'listo' else '',
            ETAPA_ABREV.get(etapa, etapa.upper()[:3]),
            ot.get('posicion',''),
            ot.get('comentario',''), ot.get('repuestos',''),
            fdate(ot.get('fecha_ingreso','')),
            fdate(ot.get('fecha_entrega','')),
            ot.get('telefono',''),
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.border = borde
            if ci == 18:  # EST
                cell.fill = fill_r
                cell.font = Font(name="Arial", bold=True, color="C0131A", size=9)
                cell.alignment = centro
            elif ci in range(12, 18):  # marcas X
                cell.fill = VERDE_F if val == 'X' else BLANCO_F
                cell.font = Font(name="Arial", bold=True,
                                 color="155724" if val == 'X' else "AAAAAA", size=9)
                cell.alignment = centro
            else:
                cell.fill = GRIS_CLAR if rn % 2 == 0 else BLANCO_F
                cell.font = Font(name="Arial", color="000000", size=9)
                cell.alignment = izq if ci in [2,20,21] else centro
        ws.row_dimensions[rn].height = 16

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────
# Debug: mostrar modo de BD
if supa:
    st.sidebar.success("☁️ BD: Supabase")
else:
    st.sidebar.info("💻 BD: Local (SQLite)")

ots = load_ots()
n_urgentes = sum(1 for o in ots if dias_restantes(o.get('fecha_entrega','')) is not None
                 and dias_restantes(o.get('fecha_entrega','')) <= 3
                 and o['etapa_actual'] != 'listo')

st.markdown(f"""
<div class='main-header'>
    <div style='font-size:28px'>🔧</div>
    <div>
        <h1>Torre de Control · FL Servital</h1>
        <p>{hoy.strftime('%A %d de %B, %Y').capitalize()} &nbsp;·&nbsp; {len(ots)} vehículos activos
        {'&nbsp;·&nbsp; <span style="color:#ff6b6b">⚠️ ' + str(n_urgentes) + ' urgentes</span>' if n_urgentes else ''}</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────
tab_hoy, tab_panel, tab_crono, tab_nueva, tab_mapa, tab_config = st.tabs([
    "📅  Hoy", "📋  Panel", "📊  Cronograma", "➕  Nueva OT", "🗺️  Mapa", "⚙️  Config"
])

# ══════════════════════════════════════════════════════
# TAB 1 — HOY
# ══════════════════════════════════════════════════════
with tab_hoy:
    st.markdown(f"### Trabajos del día — {hoy.strftime('%d/%m/%Y')}")

    # Encontrar qué hace cada operario HOY
    def trabajo_hoy(nombre_operario, tipo):
        trabajos = []
        # Etapas válidas según tipo de operario
        etapas_des = ['desabolladura', 'armado']
        etapas_pin = ['pintura', 'pulido']
        etapas_validas = etapas_des if tipo == 'des' else etapas_pin

        for ot in ots:
            prog = ot.get('programacion', {})
            operario = ot['desabollador'] if tipo == 'des' else ot['pintor']
            if operario != nombre_operario: continue
            etapa_actual = ot.get('etapa_actual', '')
            if etapa_actual == 'listo': continue

            for etapa, info in prog.items():
                if etapa not in etapas_validas: continue  # solo etapas del tipo correcto
                if not info: continue
                try:
                    fi = datetime.strptime(info['inicio'], '%Y-%m-%d').date()
                    ff = datetime.strptime(info['fin'], '%Y-%m-%d').date()
                    # Solo mostrar si la etapa está activa hoy Y no está completada
                    if fi <= hoy <= ff and not info.get('completado', False):
                        # Verificar que la etapa no fue superada según etapa_actual
                        ETAPA_IDX = {e: i for i, e in enumerate(ETAPAS)}
                        idx_actual = ETAPA_IDX.get(etapa_actual, -1)
                        idx_etapa  = ETAPA_IDX.get(etapa, -1)
                        if idx_etapa < idx_actual: continue  # etapa ya completada
                        trabajos.append({
                            'ot': ot['ot'], 'patente': ot['patente'],
                            'modelo': ot['modelo'], 'etapa': etapa,
                            'etapa_actual': etapa_actual
                        })
                except: pass
        return trabajos

    # Alertas urgentes
    urgentes = [o for o in ots if dias_restantes(o.get('fecha_entrega','')) is not None
                and dias_restantes(o.get('fecha_entrega','')) <= 3
                and o['etapa_actual'] != 'listo']
    if urgentes:
        st.markdown("#### 🚨 Urgentes")
        for u in urgentes:
            dr = dias_restantes(u['fecha_entrega'])
            color = "#c53030" if dr < 0 else "#dd6b20"
            msg = f"{'ATRASADO ' + str(abs(dr)) + ' días' if dr < 0 else 'Entrega en ' + str(dr) + ' días'}"
            st.markdown(f"""
            <div class='alerta-urgente'>
                <b>{u['ot']}</b> · {u['patente']} · {u['modelo']}
                &nbsp;{badge_etapa(u['etapa_actual'])}
                &nbsp;<span style='color:{color};font-weight:700'>{msg}</span>
                &nbsp;· Des: <b>{u['desabollador']}</b> · Pin: <b>{u['pintor']}</b>
            </div>
            """, unsafe_allow_html=True)
        st.markdown("---")

    # Tarjetas por operario
    st.markdown("#### 👷 Desabolladores")
    cols_des = st.columns(3)
    for i, des in enumerate(DESABOLLADORES):
        trabajos = trabajo_hoy(des, 'des')
        carga = len(trabajos)
        color_borde = "#48BB78" if carga == 0 else "#F6AD55" if carga <= 2 else "#FC8181"
        with cols_des[i]:
            if trabajos:
                tareas_html = "".join([
                    f"<div style='margin-bottom:6px'>"
                    f"<span style='font-family:DM Mono,monospace;font-size:12px;color:#666'>{t['ot']} · {t['patente']}</span><br>"
                    f"<b style='font-size:14px'>{t['modelo']}</b> &nbsp; {badge_etapa(t['etapa'])}"
                    f"</div>"
                    for t in trabajos
                ])
            else:
                tareas_html = "<span style='color:#aaa;font-size:13px'>Sin trabajos hoy</span>"

            st.markdown(f"""
            <div class='operario-card' style='border-color:{color_borde}'>
                <div class='operario-name'>{des}</div>
                {tareas_html}
            </div>
            """, unsafe_allow_html=True)

    st.markdown("#### 🎨 Pintores")
    cols_pin = st.columns(2)
    for i, pin in enumerate(PINTORES):
        trabajos = trabajo_hoy(pin, 'pin')
        carga = len(trabajos)
        cap = 2 if pin == 'Huaiquifil' else 1
        color_borde = "#48BB78" if carga == 0 else "#F6AD55" if carga < cap else "#FC8181"
        with cols_pin[i]:
            if trabajos:
                tareas_html = "".join([
                    f"<div style='margin-bottom:6px'>"
                    f"<span style='font-family:DM Mono,monospace;font-size:12px;color:#666'>{t['ot']} · {t['patente']}</span><br>"
                    f"<b style='font-size:14px'>{t['modelo']}</b> &nbsp; {badge_etapa(t['etapa'])}"
                    f"</div>"
                    for t in trabajos
                ])
            else:
                tareas_html = "<span style='color:#aaa;font-size:13px'>Sin trabajos hoy</span>"

            st.markdown(f"""
            <div class='operario-card' style='border-color:{color_borde}'>
                <div class='operario-name'>{pin} <span style='font-size:11px;color:#999'>(cap. {cap})</span></div>
                {tareas_html}
            </div>
            """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# TAB 2 — PANEL DE PROGRAMACIÓN
# ══════════════════════════════════════════════════════
with tab_panel:
    # Filtros rápidos
    f1, f2, f3, f4 = st.columns([2, 2, 2, 2])
    filtro_texto  = f1.text_input("🔍 Buscar patente / modelo / OT", key="p_buscar")
    filtro_etapa  = f2.selectbox("Etapa", ["Todas"] + [ETAPA_LABEL[e] for e in ETAPAS], key="p_etapa")
    filtro_des    = f3.selectbox("Desabollador", ["Todos"] + DESABOLLADORES, key="p_des")
    filtro_pin    = f4.selectbox("Pintor", ["Todos"] + PINTORES, key="p_pin")

    ots_filtradas = ots
    if filtro_texto:
        txt = filtro_texto.lower()
        ots_filtradas = [o for o in ots_filtradas if
            txt in o['patente'].lower() or txt in o['modelo'].lower() or txt in o['ot'].lower()]
    if filtro_etapa != "Todas":
        etapa_key = [k for k, v in ETAPA_LABEL.items() if v == filtro_etapa]
        if etapa_key:
            ots_filtradas = [o for o in ots_filtradas if o['etapa_actual'] == etapa_key[0]]
    if filtro_des != "Todos":
        ots_filtradas = [o for o in ots_filtradas if o['desabollador'] == filtro_des]
    if filtro_pin != "Todos":
        ots_filtradas = [o for o in ots_filtradas if o['pintor'] == filtro_pin]

    col_cap, col_dl = st.columns([3, 1])
    col_cap.caption(f"Mostrando {len(ots_filtradas)} de {len(ots)} vehículos")
    excel_buf = exportar_excel(ots)
    col_dl.download_button("📥 Descargar Excel", excel_buf,
        file_name=f"Planilla_Taller_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
    st.markdown("---")

    # Panel OT por OT
    if 'editar_ot' not in st.session_state:
        st.session_state.editar_ot = None

    for ot in ots_filtradas:
        # Fecha taller (ingreso) es la base para atrasos y urgencias
        fi_fmt = format_fecha(ot.get('fecha_ingreso',''))
        fe_fmt = format_fecha(ot.get('fecha_entrega',''))
        dr = dias_restantes(ot.get('fecha_ingreso',''))
        pct = progress_pct(ot['etapa_actual'])
        pcolor = progress_color(ot['etapa_actual'])
        es_listo = ot['etapa_actual'] == 'listo'
        tentativa = ot.get('comentario','').startswith('⚠️ Fecha tentativa')

        # Color fondo según urgencia
        if not es_listo and dr is not None and dr < 0:
            row_bg = "#fff5f5"
        elif not es_listo and dr is not None and dr <= 3:
            row_bg = "#fffbf0"
        else:
            row_bg = "white"

        # Fecha label
        if tentativa:
            fecha_label = "<span style='color:#e07b00;font-weight:600'>⚠️ Fecha tentativa</span>"
        elif fi_fmt and fi_fmt != '—':
            if not es_listo and dr is not None and dr < 0:
                fecha_label = f"<span style='color:#c53030;font-weight:700'>🚨 Taller: {fi_fmt} ({abs(dr)}d atraso)</span>"
            elif not es_listo and dr is not None and dr <= 3:
                fecha_label = f"<span style='color:#dd6b20;font-weight:700'>⚠️ Taller: {fi_fmt} ({dr}d)</span>"
            else:
                fecha_label = f"<span style='color:#444'>🔧 Taller: {fi_fmt}</span>"
            if fe_fmt and fe_fmt != '—':
                fecha_label += f" &nbsp;<span style='color:#999;font-size:11px'>· Cliente: {fe_fmt}</span>"
        else:
            fecha_label = "<span style='color:#e07b00;font-weight:600'>⚠️ Fecha tentativa</span>"

        posicion_html = f" · 📍 <b>{ot.get('posicion','')}</b>" if ot.get('posicion') else ''

        col_info, col_prog, col_acc = st.columns([3, 4, 2])

        with col_info:
            st.markdown(f"""
            <div style='background:{row_bg};padding:12px;border-radius:10px;border:1px solid #e8e8e8;height:100%'>
                <div style='font-size:11px;color:#888;font-family:DM Mono,monospace'>{ot['ot']}</div>
                <div style='font-size:16px;font-weight:700;color:#1a1a2e'>{ot['patente']} · {ot['modelo']}</div>
                <div style='font-size:12px;color:#666;margin-top:2px'>{ot.get('nombre') or '—'} · {ot.get('cia') or '—'}</div>
                <div style='margin-top:6px;font-size:12px'>{fecha_label}</div>
                <div style='font-size:12px;color:#888;margin-top:4px'>Des: <b>{ot['desabollador']}</b> · Pin: <b>{ot['pintor']}</b>{posicion_html}</div>
            </div>
            """, unsafe_allow_html=True)

        with col_prog:
            st.markdown(f"""
            <div style='background:{row_bg};padding:12px;border-radius:10px;border:1px solid #e8e8e8;height:100%'>
                <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px'>
                    <span style='font-size:12px;color:#666'>Progreso</span>
                    {badge_etapa(ot['etapa_actual'])}
                </div>
                <div class='progress-container'>
                    <div class='progress-bar' style='width:{pct}%;background:{pcolor}'></div>
                </div>
                <div style='display:flex;justify-content:space-between;margin-top:8px'>
                    {''.join([
                        f"<span style='font-size:10px;color:{"#1a1a2e" if ETAPA_PROGRESS.get(ot["etapa_actual"],0) > ETAPA_PROGRESS.get(e,0) else ("#2B6CB0" if e == ot["etapa_actual"] else "#ccc")};font-weight:{"700" if e == ot["etapa_actual"] else "400"}'>{ETAPA_LABEL[e][:3]}</span>"
                        for e in ETAPAS
                    ])}
                </div>
                {'<div style="font-size:11px;color:#888;margin-top:6px">💬 ' + ot["comentario"][:60] + ('...' if len(ot.get("comentario","")) > 60 else '') + '</div>' if ot.get("comentario") else ""}
            </div>
            """, unsafe_allow_html=True)

        with col_acc:
            sig = SIGUIENTE_ETAPA.get(ot['etapa_actual'])
            if sig:
                if st.button(f"→ {ETAPA_LABEL[sig]}", key=f"av_{ot['ot']}", use_container_width=True, type="primary"):
                    avanzar_etapa(ot['ot'], sig)
                    st.rerun()
            else:
                if st.button("🚗 Entregar", key=f"en_{ot['ot']}", use_container_width=True):
                    marcar_entregado(ot['ot'])
                    st.rerun()

            if st.button("✏️ Editar", key=f"ed_{ot['ot']}", use_container_width=True):
                st.session_state.editar_ot = ot['ot'] if st.session_state.editar_ot != ot['ot'] else None
                st.rerun()

        # Panel de edición inline
        if st.session_state.editar_ot == ot['ot']:
            with st.container():
                st.markdown(f"**✏️ Editando {ot['ot']} — {ot['patente']}**")
                e1, e2, e3 = st.columns(3)
                etapa_opts = ETAPAS
                etapa_idx = etapa_opts.index(ot['etapa_actual']) if ot['etapa_actual'] in etapa_opts else 0
                nueva_etapa = e1.selectbox("Etapa actual", etapa_opts,
                    format_func=lambda x: ETAPA_LABEL[x],
                    index=etapa_idx, key=f"et_{ot['ot']}")
                nuevo_des = e2.selectbox("Desabollador", DESABOLLADORES,
                    index=DESABOLLADORES.index(ot['desabollador']) if ot['desabollador'] in DESABOLLADORES else 0,
                    key=f"des_{ot['ot']}")
                nuevo_pin = e3.selectbox("Pintor", PINTORES,
                    index=PINTORES.index(ot['pintor']) if ot['pintor'] in PINTORES else 0,
                    key=f"pin_{ot['ot']}")
                nuevo_com = st.text_area("Comentario", value=ot.get('comentario',''), height=60, key=f"com_{ot['ot']}")
                nuevo_rep = st.text_input("Repuestos", value=ot.get('repuestos',''), key=f"rep_{ot['ot']}")
                pos_actual = ot.get('posicion','') or '— Sin asignar —'
                pos_idx = TODAS_POSICIONES.index(pos_actual) if pos_actual in TODAS_POSICIONES else 0
                nueva_pos = st.selectbox("📍 Posición en el taller", TODAS_POSICIONES,
                    index=pos_idx, key=f"pos_{ot['ot']}")

                s1, s2, _ = st.columns([1, 1, 4])
                if s1.button("💾 Guardar", key=f"save_{ot['ot']}", type="primary"):
                    pos_guardar = nueva_pos if nueva_pos != '— Sin asignar —' else ''
                    if supa:
                        supa.table("ots").update({
                            'etapa_actual': nueva_etapa, 'desabollador': nuevo_des,
                            'pintor': nuevo_pin, 'comentario': nuevo_com,
                            'repuestos': nuevo_rep, 'posicion': pos_guardar
                        }).eq("ot", ot['ot']).execute()
                    else:
                        c = conn.cursor()
                        c.execute("""UPDATE ots SET etapa_actual=?, desabollador=?, pintor=?,
                                    comentario=?, repuestos=?, posicion=? WHERE ot=?""",
                                 (nueva_etapa, nuevo_des, nuevo_pin, nuevo_com, nuevo_rep, pos_guardar, ot['ot']))
                        conn.commit()
                    st.session_state.editar_ot = None
                    st.success("✅ Guardado")
                    st.rerun()
                if s2.button("✖ Cerrar", key=f"cls_{ot['ot']}"):
                    st.session_state.editar_ot = None
                    st.rerun()

        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# TAB 3 — CRONOGRAMA
# ══════════════════════════════════════════════════════
with tab_crono:
    cr1, cr2 = st.columns([2, 3])
    vista = cr1.radio("Vista", ["Por semana", "Por desabollador", "Por pintor"], horizontal=True)
    if 'semana_offset' not in st.session_state:
        st.session_state.semana_offset = 0

    if vista == "Por semana":
        n1, n2, n3, n4 = st.columns([1, 1, 2, 2])
        if n1.button("◀ Anterior"): st.session_state.semana_offset -= 1; st.rerun()
        if n2.button("Siguiente ▶"): st.session_state.semana_offset += 1; st.rerun()
        filtro_op_crono = n3.multiselect("Filtrar operarios:",
            DESABOLLADORES + PINTORES,
            default=DESABOLLADORES + PINTORES,
            key="crono_op")

        lunes = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=st.session_state.semana_offset)
        dias = [lunes + timedelta(days=i) for i in range(5)]

        st.markdown(f"#### Semana del {lunes.strftime('%d/%m')} al {dias[-1].strftime('%d/%m/%Y')}")

        # Construir tabla
        todos_operarios = DESABOLLADORES + PINTORES
        tabla_data = {op: {d: [] for d in dias} for op in todos_operarios}

        for ot in ots:
            prog = ot.get('programacion', {})
            for etapa, info in prog.items():
                if not info: continue
                try:
                    fi = datetime.strptime(info['inicio'], '%Y-%m-%d').date()
                    ff = datetime.strptime(info['fin'], '%Y-%m-%d').date()
                except: continue
                if etapa in ('desabolladura', 'armado'):
                    op = ot['desabollador']
                elif etapa in ('pintura', 'pulido'):
                    op = ot['pintor']
                else:
                    continue
                if op not in tabla_data: continue
                for dia in dias:
                    if fi <= dia <= ff and es_laboral(dia):
                        tabla_data[op][dia].append(f"{ot['patente']} ({ETAPA_LABEL[etapa][:3]})")

        # Renderizar tabla
        cols_dias = st.columns([2] + [1]*5)
        cols_dias[0].markdown("**Operario**")
        for i, d in enumerate(dias):
            es_hoy = d == hoy
            label = f"**{'🔵 ' if es_hoy else ''}{d.strftime('%a %d/%m')}**"
            cols_dias[i+1].markdown(label)

        st.markdown("<hr style='margin:4px 0'>", unsafe_allow_html=True)

        todos_operarios_filtrados = [op for op in todos_operarios if op in filtro_op_crono]
        for op in todos_operarios_filtrados:
            cols = st.columns([2] + [1]*5)
            cols[0].markdown(f"**{op}**")
            for i, d in enumerate(dias):
                tareas = tabla_data[op][d]
                if tareas:
                    for t in tareas:
                        cols[i+1].markdown(f"<span style='font-size:12px;background:#e8f4fd;padding:2px 6px;border-radius:4px;display:block;margin-bottom:2px'>{t}</span>", unsafe_allow_html=True)
                else:
                    cols[i+1].markdown("<span style='color:#ccc;font-size:12px'>—</span>", unsafe_allow_html=True)

    elif vista == "Por desabollador":
        des_sel = st.selectbox("Seleccionar desabollador", DESABOLLADORES)
        ots_des = [o for o in ots if o['desabollador'] == des_sel]
        st.markdown(f"#### {des_sel} — {len(ots_des)} vehículos asignados")

        for ot in ots_des:
            prog = ot.get('programacion', {})
            pct = progress_pct(ot['etapa_actual'])
            pcolor = progress_color(ot['etapa_actual'])
            dr = dias_restantes(ot['fecha_entrega'])

            col_a, col_b = st.columns([2, 3])
            with col_a:
                st.markdown(f"""
                <div style='padding:10px;background:white;border-radius:8px;border:1px solid #e8e8e8'>
                    <div style='font-size:11px;color:#888'>{ot['ot']}</div>
                    <b>{ot['patente']} · {ot['modelo']}</b>
                    <div style='font-size:12px;color:#666'>Entrega: {format_fecha(ot['fecha_entrega'])}
                    {"· <span style='color:#c53030'>ATRASADO</span>" if dr is not None and dr < 0 else ""}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            with col_b:
                # Mini cronograma de etapas
                etapas_html = ""
                for e in ['desabolladura', 'armado']:
                    info = prog.get(e)
                    if info:
                        fi = format_fecha(info['inicio'])
                        ff = format_fecha(info['fin'])
                        activo = ot['etapa_actual'] == e
                        bg = "#2B6CB0" if activo else "#e8f4fd"
                        tc = "white" if activo else "#2B6CB0"
                        etapas_html += f"<span style='background:{bg};color:{tc};padding:3px 8px;border-radius:4px;font-size:11px;margin-right:4px'>{ETAPA_LABEL[e]}: {fi}→{ff}</span>"
                st.markdown(f"<div style='padding:10px'>{etapas_html}</div>", unsafe_allow_html=True)

    elif vista == "Por pintor":
        pin_sel = st.selectbox("Seleccionar pintor", PINTORES)
        ots_pin = [o for o in ots if o['pintor'] == pin_sel]
        st.markdown(f"#### {pin_sel} — {len(ots_pin)} vehículos asignados")

        for ot in ots_pin:
            prog = ot.get('programacion', {})
            dr = dias_restantes(ot['fecha_entrega'])
            col_a, col_b = st.columns([2, 3])
            with col_a:
                atras = "· <span style='color:#c53030'>ATRASADO</span>" if dr is not None and dr < 0 else ""
                st.markdown(f"<div style='padding:10px;background:white;border-radius:8px;border:1px solid #e8e8e8'>"
                    f"<div style='font-size:11px;color:#888'>{ot['ot']}</div>"
                    f"<b>{ot['patente']} · {ot['modelo']}</b>"
                    f"<div style='font-size:12px;color:#666'>Entrega: {format_fecha(ot['fecha_entrega'])} {atras}</div>"
                    f"</div>", unsafe_allow_html=True)
            with col_b:
                etapas_html = ""
                for e in ['pintura', 'pulido']:
                    info = prog.get(e)
                    if info:
                        fi = format_fecha(info['inicio'])
                        ff = format_fecha(info['fin'])
                        activo = ot['etapa_actual'] == e
                        bg = "#2B6CB0" if activo else "#e8f4fd"
                        tc = "white" if activo else "#2B6CB0"
                        etapas_html += f"<span style='background:{bg};color:{tc};padding:3px 8px;border-radius:4px;font-size:11px;margin-right:4px'>{ETAPA_LABEL[e]}: {fi}→{ff}</span>"
                st.markdown(f"<div style='padding:10px'>{etapas_html}</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# TAB 4 — NUEVA OT
# ══════════════════════════════════════════════════════
with tab_nueva:
    st.markdown("### ➕ Ingresar nuevo vehículo")
    st.markdown("El sistema asignará automáticamente el equipo más libre.")

    with st.form("form_nueva_ot"):
        st.markdown("**Datos del vehículo**")
        n1, n2, n3, n4 = st.columns(4)
        f_ot      = n1.text_input("N° OT *", placeholder="219999")
        f_patente = n2.text_input("Patente *", placeholder="XXXX99")
        f_modelo  = n3.text_input("Modelo *", placeholder="Tiggo 8")
        f_color   = n4.text_input("Color")

        n5, n6, n7 = st.columns(3)
        f_dano    = n5.selectbox("Tipo de daño", ["MEDIANO", "LEVE", "GRAVE"])
        f_ingreso = n6.date_input("Fecha ingreso", value=hoy)
        f_entrega = n7.date_input("Fecha entrega estimada", value=hoy + timedelta(days=14))

        st.markdown("**Datos del cliente**")
        c1, c2, c3, c4 = st.columns(4)
        f_nombre  = c1.text_input("Nombre cliente")
        f_cia     = c2.text_input("Compañía (CIA)")
        f_ase     = c3.text_input("ASE")
        f_tel     = c4.text_input("Teléfono")

        st.markdown("**Asignación de equipo**")
        a1, a2, a3 = st.columns([1, 1, 2])
        auto_asignar = a1.checkbox("🤖 Asignación automática", value=True)

        ots_actual = load_ots()
        if auto_asignar:
            sug_des = mejor_desabollador(ots_actual, f_dano)
            sug_pin = mejor_pintor(ots_actual, f_dano)
            a2.info(f"Se asignará:\n**Des:** {sug_des}\n**Pin:** {sug_pin}")
            f_des = sug_des
            f_pin = sug_pin
        else:
            f_des = a2.selectbox("Desabollador", DESABOLLADORES)
            f_pin = a3.selectbox("Pintor", PINTORES)

        f_comentario = st.text_area("Comentario / observaciones", height=60)
        f_repuestos  = st.text_input("Estado repuestos")

        submitted = st.form_submit_button("🚀 Programar vehículo", type="primary", use_container_width=True)

        if submitted:
            if not f_ot or not f_patente or not f_modelo:
                st.error("OT, Patente y Modelo son obligatorios")
            else:
                ot_id = f"OT-{f_ot.strip()}"
                ots_ids = {o['ot'] for o in ots_actual}
                if ot_id in ots_ids:
                    st.error(f"{ot_id} ya existe en el sistema")
                else:
                    prog = programar_ot(f_des, f_pin, f_dano, f_ingreso, ots_actual)
                    data = {
                        'ot': ot_id, 'nombre': f_nombre, 'cia': f_cia, 'ase': f_ase,
                        'modelo': f_modelo, 'color': f_color, 'patente': f_patente,
                        'dano': f_dano, 'desabollador': f_des, 'pintor': f_pin,
                        'etapa_actual': 'desabolladura',
                        'fecha_ingreso': f_ingreso.strftime('%Y-%m-%d'),
                        'fecha_entrega': f_entrega.strftime('%Y-%m-%d'),
                        'comentario': f_comentario, 'repuestos': f_repuestos, 'telefono': f_tel,
                        'programacion': prog,
                    }
                    save_ot(data)
                    st.success(f"✅ {ot_id} programado correctamente — Desabollador: {f_des} · Pintor: {f_pin}")
                    st.rerun()

# ══════════════════════════════════════════════════════
# TAB 5 — MAPA
# ══════════════════════════════════════════════════════
with tab_mapa:
    import streamlit.components.v1 as components

    st.markdown("### 🗺️ Mapa del Taller")
    st.caption("Clic en cualquier posición para asignar o mover vehículos. Los cambios se guardan automáticamente.")

    # Preparar datos JSON de OTs
    ots_json_mapa = json.dumps([{
        'ot': o['ot'], 'patente': o['patente'], 'modelo': o['modelo'],
        'color': o.get('color',''), 'cia': o.get('cia',''),
        'etapa': o.get('etapa_actual',''), 'desabollador': o.get('desabollador',''),
        'pintor': o.get('pintor',''), 'posicion': o.get('posicion',''),
        'comentario': o.get('comentario','')
    } for o in ots])

    SUPA_URL = "https://ptmnntxjptgzczlvixzh.supabase.co"
    SUPA_KEY_W = "sb_secret_8efCxndPaRABCrxOL5e1cw_TKUURWX6"

    mapa_html = """<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:system-ui,sans-serif;background:#f0f2f5}
#toolbar{display:flex;gap:8px;align-items:center;padding:8px 12px;background:#1a1a2e;color:white;font-size:12px;flex-wrap:wrap}
#toolbar select{padding:4px 8px;border-radius:6px;border:none;background:#2d2d4e;color:white;font-size:12px}
#status{margin-left:auto;font-size:11px;opacity:.8}
#map-wrap{width:100%;overflow:auto;background:#e8e8e8;padding:8px}
.zg{cursor:pointer;transition:.15s}
.zg:hover rect{opacity:.8}
.zg.ocupado rect{stroke:#ffd700;stroke-width:2.5}
.zg.ocupado-multi rect{stroke:#ff4444;stroke-width:2.5}
#side-panel{position:fixed;right:-380px;top:0;width:360px;height:100%;background:white;
  box-shadow:-4px 0 20px rgba(0,0,0,.2);transition:.3s;z-index:999;overflow-y:auto;padding:16px}
#side-panel.open{right:0}
#ph{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px}
#pt{font-size:15px;font-weight:700;color:#1a1a2e}
#cp{background:none;border:none;font-size:22px;cursor:pointer;color:#888}
.ot-card{background:#f8f9fa;border-radius:10px;padding:12px;margin-bottom:10px;border:1px solid #e8e8e8}
.ot-pat{font-size:16px;font-weight:700;color:#1a1a2e}
.ot-sub{font-size:12px;color:#666;margin:2px 0 6px}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.btn-blue{padding:6px;border:1px solid #2B6CB0;background:white;color:#2B6CB0;border-radius:6px;cursor:pointer;font-size:11px;flex:1}
.btn-red{padding:6px;border:1px solid #e53e3e;background:white;color:#e53e3e;border-radius:6px;cursor:pointer;font-size:11px;flex:1}
.sep{font-size:11px;color:#888;text-transform:uppercase;letter-spacing:.5px;margin:10px 0 6px}
select.ot-sel{width:100%;padding:8px;border:1px solid #ddd;border-radius:8px;font-size:12px;margin-bottom:8px}
.btn-assign{width:100%;padding:10px;background:#1a1a2e;color:white;border:none;border-radius:8px;cursor:pointer;font-weight:600;font-size:13px}
.modo-mover{background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:8px 12px;font-size:12px;margin-bottom:10px}
</style>
</head><body>
<div id="toolbar">
  <b style="font-size:13px">🗺️ Taller D&P</b>
  <select id="fil-etapa" onchange="filtrar()">
    <option value="">Todas las etapas</option>
    <option value="desabolladura">Desabolladura</option>
    <option value="pintura">Pintura</option>
    <option value="armado">Armado</option>
    <option value="pulido">Pulido</option>
    <option value="terminaciones">Terminaciones</option>
    <option value="lavado">Lavado</option>
  </select>
  <select id="fil-des" onchange="filtrar()">
    <option value="">Todos desabolladores</option>
    <option value="Rojas">Rojas</option>
    <option value="González">González</option>
    <option value="Carvajal">Carvajal</option>
  </select>
  <select id="fil-pin" onchange="filtrar()">
    <option value="">Todos pintores</option>
    <option value="Percy">Percy</option>
    <option value="Huaiquifil">Huaiquifil</option>
  </select>
  <span id="status">Cargando...</span>
</div>
<div id="map-wrap">
<svg viewBox="0 0 1400 960" xmlns="http://www.w3.org/2000/svg" style="width:100%;min-width:1000px;background:#dde0e6">

<!-- ZONA CARROCERÍA -->
<rect x="10" y="10" width="240" height="470" rx="8" fill="#c8d8ea" stroke="#9ab" stroke-width="1"/>
<text x="130" y="28" text-anchor="middle" font-size="11" fill="#336" font-weight="700">ZONA CARROCERÍA (Des/Arm)</text>
<g id="carr1"  class="zg" onclick="clic('carr1')"><rect  x="18" y="35"  width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="51"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr1</text></g>
<g id="carr2"  class="zg" onclick="clic('carr2')"><rect  x="18" y="67"  width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="83"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr2</text></g>
<g id="carr3"  class="zg" onclick="clic('carr3')"><rect  x="18" y="99"  width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="115" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr3</text></g>
<g id="carr4"  class="zg" onclick="clic('carr4')"><rect  x="18" y="131" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="147" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr4</text></g>
<g id="carr5"  class="zg" onclick="clic('carr5')"><rect  x="18" y="210" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="226" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr5</text></g>
<g id="carr6"  class="zg" onclick="clic('carr6')"><rect  x="18" y="242" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="258" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr6</text></g>
<g id="carr7"  class="zg" onclick="clic('carr7')"><rect  x="18" y="274" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="290" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr7</text></g>
<g id="carr8"  class="zg" onclick="clic('carr8')"><rect  x="18" y="306" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="322" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr8</text></g>
<g id="carr9"  class="zg" onclick="clic('carr9')"><rect  x="18" y="338" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="354" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr9</text></g>
<g id="carr10" class="zg" onclick="clic('carr10')"><rect x="18" y="370" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="68"  y="386" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr10</text></g>
<g id="carr11" class="zg" onclick="clic('carr11')"><rect x="130" y="210" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="180" y="226" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr11</text></g>
<g id="carr12" class="zg" onclick="clic('carr12')"><rect x="130" y="242" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="180" y="258" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr12</text></g>
<g id="carr13" class="zg" onclick="clic('carr13')"><rect x="130" y="274" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="180" y="290" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr13</text></g>
<g id="carr14" class="zg" onclick="clic('carr14')"><rect x="130" y="306" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="180" y="322" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr14</text></g>
<g id="carr15" class="zg" onclick="clic('carr15')"><rect x="130" y="338" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="180" y="354" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr15</text></g>
<g id="carr16" class="zg" onclick="clic('carr16')"><rect x="130" y="370" width="100" height="28" rx="4" fill="#5B8CBA"/><text x="180" y="386" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Carr16</text></g>
<!-- Pre Entrega -->
<g id="pe1" class="zg" onclick="clic('pe1')"><rect x="130" y="131" width="48" height="28" rx="4" fill="#7BAF6E"/><text x="154" y="147" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">PE1</text></g>
<g id="pe2" class="zg" onclick="clic('pe2')"><rect x="182" y="131" width="48" height="28" rx="4" fill="#7BAF6E"/><text x="206" y="147" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">PE2</text></g>
<g id="pe3" class="zg" onclick="clic('pe3')"><rect x="130" y="99"  width="48" height="28" rx="4" fill="#7BAF6E"/><text x="154" y="115" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">PE3</text></g>
<g id="pe4" class="zg" onclick="clic('pe4')"><rect x="182" y="99"  width="48" height="28" rx="4" fill="#7BAF6E"/><text x="206" y="115" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">PE4</text></g>
<!-- Lavado -->
<g id="lav1" class="zg" onclick="clic('lav1')"><rect x="18"  y="430" width="105" height="38" rx="4" fill="#4A7FB5"/><text x="70"  y="449" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Lavado 1</text></g>
<g id="lav2" class="zg" onclick="clic('lav2')"><rect x="130" y="430" width="105" height="38" rx="4" fill="#4A7FB5"/><text x="182" y="449" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Lavado 2</text></g>

<!-- ESP ARMADO -->
<rect x="260" y="200" width="120" height="270" rx="8" fill="#e8d5ff" stroke="#9B59B6" stroke-width="1"/>
<text x="320" y="218" text-anchor="middle" font-size="10" fill="#4a0080" font-weight="700">ESP.ARMADO</text>
<g id="est11" class="zg" onclick="clic('est11')"><rect x="268" y="225" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="238" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est11</text></g>
<g id="est12" class="zg" onclick="clic('est12')"><rect x="268" y="255" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="268" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est12</text></g>
<g id="est13" class="zg" onclick="clic('est13')"><rect x="268" y="285" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="298" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est13</text></g>
<g id="est14" class="zg" onclick="clic('est14')"><rect x="268" y="315" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="328" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est14</text></g>
<g id="est15" class="zg" onclick="clic('est15')"><rect x="268" y="345" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="358" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est15</text></g>
<g id="est16" class="zg" onclick="clic('est16')"><rect x="268" y="375" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="388" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est16</text></g>
<g id="est17" class="zg" onclick="clic('est17')"><rect x="268" y="405" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="418" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est17</text></g>
<g id="est18" class="zg" onclick="clic('est18')"><rect x="268" y="435" width="104" height="26" rx="4" fill="#8B6BB0"/><text x="320" y="448" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est18</text></g>

<!-- ESPERA PINTURA P1-P15 -->
<rect x="395" y="10" width="270" height="420" rx="8" fill="#f8d7d7" stroke="#C0131A" stroke-width="1.5" stroke-dasharray="5"/>
<text x="530" y="28" text-anchor="middle" font-size="11" fill="#C0131A" font-weight="700">ESPERA INGRESO PINTURA</text>
<g id="p5"  class="zg" onclick="clic('p5')"><rect  x="405" y="35"  width="78" height="28" rx="4" fill="#C0131A"/><text x="444" y="51"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P5</text></g>
<g id="p10" class="zg" onclick="clic('p10')"><rect x="490" y="35"  width="78" height="28" rx="4" fill="#C0131A"/><text x="529" y="51"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P10</text></g>
<g id="p15" class="zg" onclick="clic('p15')"><rect x="575" y="35"  width="78" height="28" rx="4" fill="#C0131A"/><text x="614" y="51"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P15</text></g>
<g id="p4"  class="zg" onclick="clic('p4')"><rect  x="405" y="67"  width="78" height="28" rx="4" fill="#C0131A"/><text x="444" y="83"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P4</text></g>
<g id="p9"  class="zg" onclick="clic('p9')"><rect  x="490" y="67"  width="78" height="28" rx="4" fill="#C0131A"/><text x="529" y="83"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P9</text></g>
<g id="p14" class="zg" onclick="clic('p14')"><rect x="575" y="67"  width="78" height="28" rx="4" fill="#C0131A"/><text x="614" y="83"  text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P14</text></g>
<g id="p3"  class="zg" onclick="clic('p3')"><rect  x="405" y="99"  width="78" height="28" rx="4" fill="#C0131A"/><text x="444" y="115" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P3</text></g>
<g id="p8"  class="zg" onclick="clic('p8')"><rect  x="490" y="99"  width="78" height="28" rx="4" fill="#C0131A"/><text x="529" y="115" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P8</text></g>
<g id="p13" class="zg" onclick="clic('p13')"><rect x="575" y="99"  width="78" height="28" rx="4" fill="#C0131A"/><text x="614" y="115" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P13</text></g>
<g id="p2"  class="zg" onclick="clic('p2')"><rect  x="405" y="131" width="78" height="28" rx="4" fill="#C0131A"/><text x="444" y="147" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P2</text></g>
<g id="p7"  class="zg" onclick="clic('p7')"><rect  x="490" y="131" width="78" height="28" rx="4" fill="#C0131A"/><text x="529" y="147" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P7</text></g>
<g id="p12" class="zg" onclick="clic('p12')"><rect x="575" y="131" width="78" height="28" rx="4" fill="#C0131A"/><text x="614" y="147" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P12</text></g>
<g id="p1"  class="zg" onclick="clic('p1')"><rect  x="405" y="163" width="78" height="28" rx="4" fill="#C0131A"/><text x="444" y="179" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P1</text></g>
<g id="p6"  class="zg" onclick="clic('p6')"><rect  x="490" y="163" width="78" height="28" rx="4" fill="#C0131A"/><text x="529" y="179" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P6</text></g>
<g id="p11" class="zg" onclick="clic('p11')"><rect x="575" y="163" width="78" height="28" rx="4" fill="#C0131A"/><text x="614" y="179" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">P11</text></g>

<!-- ESP PULIDO Est19-30 -->
<rect x="680" y="430" width="150" height="400" rx="8" fill="#ddeeff" stroke="#4A7FB5" stroke-width="1"/>
<text x="755" y="448" text-anchor="middle" font-size="10" fill="#004085" font-weight="700">ESP.PULIDO</text>
<g id="est19" class="zg" onclick="clic('est19')"><rect x="688" y="455" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="468" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est19</text></g>
<g id="est20" class="zg" onclick="clic('est20')"><rect x="688" y="485" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="498" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est20</text></g>
<g id="est21" class="zg" onclick="clic('est21')"><rect x="688" y="515" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="528" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est21</text></g>
<g id="est22" class="zg" onclick="clic('est22')"><rect x="688" y="545" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="558" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est22</text></g>
<g id="est23" class="zg" onclick="clic('est23')"><rect x="688" y="575" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="588" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est23</text></g>
<g id="est24" class="zg" onclick="clic('est24')"><rect x="688" y="605" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="618" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est24</text></g>
<g id="est25" class="zg" onclick="clic('est25')"><rect x="688" y="635" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="648" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est25</text></g>
<g id="est26" class="zg" onclick="clic('est26')"><rect x="688" y="665" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="678" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est26</text></g>
<g id="est27" class="zg" onclick="clic('est27')"><rect x="688" y="695" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="708" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est27</text></g>
<g id="est28" class="zg" onclick="clic('est28')"><rect x="688" y="725" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="738" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est28</text></g>
<g id="est29" class="zg" onclick="clic('est29')"><rect x="688" y="755" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="768" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est29</text></g>
<g id="est30" class="zg" onclick="clic('est30')"><rect x="688" y="785" width="134" height="26" rx="4" fill="#4A7FB5"/><text x="755" y="798" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Est30</text></g>

<!-- PINTURA Pint1-16 -->
<rect x="840" y="430" width="90" height="500" rx="8" fill="#ffd5d5" stroke="#C0131A" stroke-width="1"/>
<text x="885" y="448" text-anchor="middle" font-size="10" fill="#C0131A" font-weight="700">PINTURA</text>
<g id="pint1"  class="zg" onclick="clic('pint1')"><rect  x="848" y="455" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="468" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint1</text></g>
<g id="pint2"  class="zg" onclick="clic('pint2')"><rect  x="848" y="485" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="498" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint2</text></g>
<g id="pint3"  class="zg" onclick="clic('pint3')"><rect  x="848" y="515" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="528" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint3</text></g>
<g id="pint4"  class="zg" onclick="clic('pint4')"><rect  x="848" y="545" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="558" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint4</text></g>
<g id="pint5"  class="zg" onclick="clic('pint5')"><rect  x="848" y="575" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="588" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint5</text></g>
<g id="pint6"  class="zg" onclick="clic('pint6')"><rect  x="848" y="605" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="618" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint6</text></g>
<g id="pint7"  class="zg" onclick="clic('pint7')"><rect  x="848" y="635" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="648" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint7</text></g>
<g id="pint8"  class="zg" onclick="clic('pint8')"><rect  x="848" y="665" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="678" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint8</text></g>
<g id="pint9"  class="zg" onclick="clic('pint9')"><rect  x="848" y="695" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="708" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint9</text></g>
<g id="pint10" class="zg" onclick="clic('pint10')"><rect x="848" y="725" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="738" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint10</text></g>
<g id="pint11" class="zg" onclick="clic('pint11')"><rect x="848" y="755" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="768" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint11</text></g>
<g id="pint12" class="zg" onclick="clic('pint12')"><rect x="848" y="785" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="798" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint12</text></g>
<g id="pint13" class="zg" onclick="clic('pint13')"><rect x="848" y="815" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="828" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint13</text></g>
<g id="pint14" class="zg" onclick="clic('pint14')"><rect x="848" y="845" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="858" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint14</text></g>
<g id="pint15" class="zg" onclick="clic('pint15')"><rect x="848" y="875" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="888" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint15</text></g>
<g id="pint16" class="zg" onclick="clic('pint16')"><rect x="848" y="905" width="74" height="26" rx="4" fill="#C0131A"/><text x="885" y="918" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pint16</text></g>

<!-- HORNOS -->
<rect x="950" y="430" width="90" height="160" rx="8" fill="#d4f5ef" stroke="#3D9E8C" stroke-width="1"/>
<text x="995" y="448" text-anchor="middle" font-size="10" fill="#1a5c52" font-weight="700">HORNOS</text>
<g id="h1" class="zg" onclick="clic('h1')"><rect x="958" y="455" width="74" height="40" rx="4" fill="#3D9E8C"/><text x="995" y="475" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pintura H1</text></g>
<g id="h2" class="zg" onclick="clic('h2')"><rect x="958" y="500" width="74" height="40" rx="4" fill="#3D9E8C"/><text x="995" y="520" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pintura H2</text></g>
<g id="h3" class="zg" onclick="clic('h3')"><rect x="958" y="545" width="74" height="40" rx="4" fill="#3D9E8C"/><text x="995" y="565" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pintura H3</text></g>

<!-- PULIDO Lav3 -->
<g id="lav3" class="zg" onclick="clic('lav3')"><rect x="950" y="610" width="90" height="40" rx="4" fill="#B794F4"/><text x="995" y="630" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Pulido (Lav3)</text></g>

<!-- MECÁNICA -->
<rect x="1060" y="160" width="220" height="100" rx="8" fill="#e0e0e0" stroke="#888" stroke-width="1"/>
<text x="1170" y="178" text-anchor="middle" font-size="10" fill="#444" font-weight="700">MECÁNICA</text>
<g id="mec1" class="zg" onclick="clic('mec1')"><rect x="1068" y="185" width="60" height="32" rx="4" fill="#888"/><text x="1098" y="201" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Mec1</text></g>
<g id="mec2" class="zg" onclick="clic('mec2')"><rect x="1138" y="185" width="60" height="32" rx="4" fill="#888"/><text x="1168" y="201" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Mec2</text></g>
<g id="mec3" class="zg" onclick="clic('mec3')"><rect x="1208" y="185" width="60" height="32" rx="4" fill="#888"/><text x="1238" y="201" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Mec3</text></g>
<g id="mec4" class="zg" onclick="clic('mec4')"><rect x="1068" y="222" width="60" height="28" rx="4" fill="#aaa"/><text x="1098" y="236" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Mec4</text></g>
<g id="mec5" class="zg" onclick="clic('mec5')"><rect x="1138" y="222" width="60" height="28" rx="4" fill="#aaa"/><text x="1168" y="236" text-anchor="middle" dominant-baseline="central" font-size="9" fill="#fff">Mec5</text></g>

<!-- LEYENDA -->
<rect x="1060" y="280" width="260" height="140" rx="8" fill="white" stroke="#ddd" stroke-width="1" fill-opacity=".95"/>
<text x="1075" y="300" font-size="11" fill="#333" font-weight="700">Leyenda</text>
<rect x="1075" y="308" width="12" height="12" rx="2" fill="#5B8CBA"/><text x="1093" y="318" font-size="9" fill="#333">Carrocería (Des/Arm)</text>
<rect x="1075" y="325" width="12" height="12" rx="2" fill="#C0131A"/><text x="1093" y="335" font-size="9" fill="#333">Pintura (Espera y Ejecución)</text>
<rect x="1075" y="342" width="12" height="12" rx="2" fill="#3D9E8C"/><text x="1093" y="352" font-size="9" fill="#333">Hornos (Pintura en curso)</text>
<rect x="1075" y="359" width="12" height="12" rx="2" fill="#B794F4"/><text x="1093" y="369" font-size="9" fill="#333">Pulido</text>
<rect x="1075" y="376" width="12" height="12" rx="2" fill="#4A7FB5"/><text x="1093" y="386" font-size="9" fill="#333">Lavado / Esp. Pulido</text>
<rect x="1075" y="393" width="12" height="12" rx="2" fill="#7BAF6E"/><text x="1093" y="403" font-size="9" fill="#333">Pre Entrega / Control Calidad</text>
<rect x="1190" y="308" width="12" height="12" rx="2" fill="none" stroke="#ffd700" stroke-width="2"/><text x="1208" y="318" font-size="9" fill="#333">1 OT</text>
<rect x="1190" y="325" width="12" height="12" rx="2" fill="none" stroke="#ff4444" stroke-width="2"/><text x="1208" y="335" font-size="9" fill="#333">2+ OTs</text>

</svg>
</div>

<!-- Panel lateral -->
<div id="side-panel">
  <div id="ph"><span id="pt">—</span><button id="cp" onclick="cerrar()">✕</button></div>
  <div id="pb"></div>
</div>

<script>
const OTS = __OTS_JSON__;
const SB_URL = "__SB_URL__";
const SB_KEY = "__SB_KEY__";

const EC = {desabolladura:'#F6AD55',pintura:'#FC8181',armado:'#F6AD55',pulido:'#B794F4',terminaciones:'#68D391',lavado:'#4FD1C5',listo:'#48BB78'};
const NOMBRES = {carr:'Carrocería',p:'Espera Pintura',est:'Est',lav:'Lavado',pe:'Pre Entrega',h:'Horno',pint:'Pintura',mec:'Mecánica'};

let pm = {};
function rebuildPM() {
  pm = {};
  OTS.forEach(o => {
    if (!o.posicion) return;
    const k = o.posicion.toLowerCase();
    if (!pm[k]) pm[k] = [];
    pm[k].push(o);
  });
}
rebuildPM();

function pintar() {
  document.querySelectorAll('.zg').forEach(g => {
    const ots = pm[g.id.toLowerCase()] || [];
    g.classList.remove('ocupado','ocupado-multi');
    g.querySelectorAll('.badge-cnt').forEach(b=>b.remove());
    if (!ots.length) return;
    g.classList.add(ots.length===1?'ocupado':'ocupado-multi');
    const txt = g.querySelector('text');
    if (txt) {
      const b = document.createElementNS('http://www.w3.org/2000/svg','text');
      b.setAttribute('x', parseFloat(txt.getAttribute('x'))+16);
      b.setAttribute('y', parseFloat(txt.getAttribute('y'))-6);
      b.setAttribute('text-anchor','middle');b.setAttribute('font-size','7');
      b.setAttribute('font-weight','bold');b.setAttribute('class','badge-cnt');
      b.setAttribute('fill', ots.length===1?'#ffd700':'#ff4444');
      b.textContent = ots.length>1?ots.length:'●';
      g.appendChild(b);
    }
  });
}

function filtrar() {
  const et=document.getElementById('fil-etapa').value;
  const de=document.getElementById('fil-des').value;
  const pi=document.getElementById('fil-pin').value;
  document.querySelectorAll('.zg').forEach(g => {
    const ots=pm[g.id.toLowerCase()]||[];
    const match=ots.some(o=>(!et||o.etapa===et)&&(!de||o.desabollador===de)&&(!pi||o.pintor===pi));
    g.style.opacity=(ots.length===0||match)?'1':'0.25';
  });
}

let modoMover=null;
function clic(id) {
  if (modoMover) {
    guardar(modoMover,id);
    modoMover=null;
    document.querySelectorAll('.zg').forEach(g=>g.style.outline='');
    document.getElementById('status').textContent='✓ Movido';
    return;
  }
  const ots=pm[id.toLowerCase()]||[];
  const prefix=id.replace(/[0-9]+[a-z]?$/,'');
  const nombre=(NOMBRES[prefix]||id.toUpperCase())+' '+id.replace(/[a-zA-Z]+/g,'');
  document.getElementById('pt').textContent=nombre;
  let html='';
  if (ots.length) {
    html+='<div>';
    ots.forEach(o=>{
      const c=EC[o.etapa]||'#ccc';
      html+=`<div class="ot-card">
        <div style="display:flex;justify-content:space-between;align-items:center">
          <span class="ot-pat">${o.patente}</span>
          <span class="badge" style="background:${c};color:#1a1a2e">${o.etapa}</span>
        </div>
        <div class="ot-sub">${o.ot} · ${o.modelo}</div>
        <div class="ot-sub">Des: ${o.desabollador} · Pin: ${o.pintor}</div>
        <div style="display:flex;gap:6px;margin-top:8px">
          <button class="btn-blue" onclick="iniciarMover('${o.ot}','${o.patente}')">📦 Mover</button>
          <button class="btn-red"  onclick="quitar('${o.ot}')">✕ Quitar</button>
        </div>
      </div>`;
    });
    html+='</div>';
  }
  const disp=OTS.filter(o=>!o.posicion||o.posicion.toLowerCase()!==id.toLowerCase());
  html+=`<div class="sep">Asignar OT</div>
  <select class="ot-sel" id="sel-ot">
    <option value="">— Seleccionar —</option>
    ${disp.map(o=>`<option value="${o.ot}">${o.patente} · ${o.modelo}${o.posicion?' (en '+o.posicion+')':''}</option>`).join('')}
  </select>
  <button class="btn-assign" onclick="asignar('${id}')">✅ Asignar aquí</button>`;
  document.getElementById('pb').innerHTML=html;
  document.getElementById('side-panel').classList.add('open');
}

function cerrar(){document.getElementById('side-panel').classList.remove('open');}
function asignar(posId){
  const v=document.getElementById('sel-ot').value;
  if(!v){alert('Selecciona una OT');return;}
  guardar(v,posId);
}
function quitar(otId){if(confirm('¿Quitar posición?'))guardar(otId,'');}
function iniciarMover(otId,pat){
  modoMover=otId;cerrar();
  document.getElementById('status').textContent=`📦 Clic en nueva posición para ${pat}...`;
  document.querySelectorAll('.zg').forEach(g=>g.style.outline='2px dashed #2B6CB0');
}
async function guardar(otId,posId){
  const posNorm=posId?posId.charAt(0).toUpperCase()+posId.slice(1):'';
  try{
    const r=await fetch(`${SB_URL}/rest/v1/ots?ot=eq.${encodeURIComponent(otId)}`,{
      method:'PATCH',
      headers:{"apikey":SB_KEY,"Authorization":"Bearer "+SB_KEY,"Content-Type":"application/json","Prefer":"return=minimal"},
      body:JSON.stringify({posicion:posNorm})
    });
    if(!r.ok)throw new Error('HTTP '+r.status);
    OTS.forEach(o=>{if(o.ot===otId)o.posicion=posNorm;});
    rebuildPM();pintar();cerrar();
    document.getElementById('status').textContent='✓ Guardado';
  }catch(e){alert('Error: '+e.message);}
}
pintar();
document.getElementById('status').textContent='""" + str(len(ots)) + """ OTs cargadas';
</script>
</body></html>"""

    # Inject data
    mapa_html = mapa_html.replace('__OTS_JSON__', ots_json_mapa)
    mapa_html = mapa_html.replace('__SB_URL__', SUPA_URL)
    mapa_html = mapa_html.replace('__SB_KEY__', SUPA_KEY_W)

    components.html(mapa_html, height=750, scrolling=True)

# ══════════════════════════════════════════════════════
# TAB 6 — CONFIG
# ══════════════════════════════════════════════════════
with tab_config:
    st.markdown("### ⚙️ Configuración")

    # Panel de pesos de asignación
    with st.expander("⚖️ Pesos de asignación automática", expanded=False):
        st.caption("Define qué porcentaje de OTs se asigna a cada operario según el tipo de daño. El sistema respeta estos pesos al asignar automáticamente.")
        pesos = get_pesos()
        DANOS = ['LEVE', 'MEDIANO', 'GRAVE']

        st.markdown("**Desabolladores**")
        for dano in DANOS:
            st.markdown(f"*{dano}*")
            cols = st.columns(len(DESABOLLADORES))
            total = 0
            nuevos = {}
            for i, des in enumerate(DESABOLLADORES):
                val = cols[i].number_input(
                    des, min_value=0, max_value=100,
                    value=int(pesos['desabolladores'][dano].get(des, 33)),
                    key=f"peso_des_{dano}_{des}", step=5
                )
                nuevos[des] = val
                total += val
            color = "🟢" if total == 100 else "🔴"
            st.caption(f"{color} Total: {total}% {'✓' if total == 100 else '— debe sumar 100%'}")
            pesos['desabolladores'][dano] = nuevos

        st.markdown("---")
        st.markdown("**Pintores**")
        for dano in DANOS:
            st.markdown(f"*{dano}*")
            cols = st.columns(len(PINTORES))
            total = 0
            nuevos = {}
            for i, pin in enumerate(PINTORES):
                val = cols[i].number_input(
                    pin, min_value=0, max_value=100,
                    value=int(pesos['pintores'][dano].get(pin, 50)),
                    key=f"peso_pin_{dano}_{pin}", step=5
                )
                nuevos[pin] = val
                total += val
            color = "🟢" if total == 100 else "🔴"
            st.caption(f"{color} Total: {total}% {'✓' if total == 100 else '— debe sumar 100%'}")
            pesos['pintores'][dano] = nuevos

        if st.button("💾 Guardar pesos", type="primary", key="btn_guardar_pesos"):
            save_pesos(pesos)
            st.success("✅ Pesos guardados correctamente")
        st.info("💡 Los pesos se aplican en la próxima asignación automática.")

    st.markdown("---")
    # Importar planilla
    with st.expander("📂 Importar desde planilla Excel", expanded=True):
        st.caption("Sube la planilla del taller para cargar los vehículos en trabajo.")
        archivo = st.file_uploader("Seleccionar archivo Excel", type=["xlsx", "xls"])
        if archivo:
            try:
                # Intentar leer hoja 'EN TRABAJO', si no existe usar la primera hoja
                xl = pd.ExcelFile(archivo)
                sheet = 'EN TRABAJO' if 'EN TRABAJO' in xl.sheet_names else xl.sheet_names[0]
                # Detectar fila de encabezado automáticamente
                df_test = pd.read_excel(archivo, sheet_name=sheet, skiprows=0, nrows=3)
                if 'O T' in df_test.columns or 'PATENTE' in df_test.columns:
                    skip = 0
                else:
                    skip = 1
                df_raw = pd.read_excel(archivo, sheet_name=sheet, skiprows=skip)
                df_raw.columns = [str(c).strip() for c in df_raw.columns]
                st.success(f"✅ {len(df_raw)} filas encontradas")
                st.dataframe(df_raw[['O T','MODELO','PATENTE','DES','PIN','EST']].head(5),
                           hide_index=True, use_container_width=True)
                if st.button("📥 Importar al sistema", type="primary"):
                    n = importar_planilla(df_raw)
                    st.success(f"✅ {n} vehículos importados")
                    st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")

    # Estadísticas
    with st.expander("📊 Resumen del sistema"):
        total = len(ots)
        por_etapa = Counter(o['etapa_actual'] for o in ots)
        por_des   = Counter(o['desabollador'] for o in ots)
        por_pin   = Counter(o['pintor'] for o in ots)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total activos", total)
        m2.metric("Urgentes (≤3 días)", n_urgentes)
        m3.metric("Listos para entrega", por_etapa.get('listo', 0))
        m4.metric("En desabolladura", por_etapa.get('desabolladura', 0))

        st.markdown("**Por etapa:**")
        for e in ETAPAS:
            n = por_etapa.get(e, 0)
            st.markdown(f"- {ETAPA_LABEL[e]}: **{n}** vehículos")

    # Peligro
    with st.expander("⚠️ Zona de peligro"):
        st.warning("Esta acción elimina TODOS los vehículos del sistema y no se puede deshacer.")
        if 'confirmar_limpiar' not in st.session_state:
            st.session_state.confirmar_limpiar = False
        if not st.session_state.confirmar_limpiar:
            if st.button("🗑️ Limpiar toda la planilla", type="secondary"):
                st.session_state.confirmar_limpiar = True
                st.rerun()
        else:
            st.error("¿Estás seguro? Esta acción no se puede deshacer.")
            c1, c2 = st.columns(2)
            if c1.button("✅ Sí, eliminar todo", type="primary"):
                if supa:
                    supa.table("ots").delete().neq("ot", "").execute()
                else:
                    c = conn.cursor()
                    c.execute("DELETE FROM ots")
                    conn.commit()
                st.session_state.confirmar_limpiar = False
                st.success("Planilla limpiada correctamente")
                st.rerun()
            if c2.button("❌ Cancelar"):
                st.session_state.confirmar_limpiar = False
                st.rerun()
